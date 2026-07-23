import glob
import os
from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from src.metrics.cer import find_best_ground_truth_match, compute_average_cer

def build_xlsx_report(detections_df: pd.DataFrame, ground_truth_plates=None) -> BytesIO:
    """
    Generates a stylized Excel report with embedded crop images of vehicles and license plates.
    
    If ground_truth_plates is provided, adds CER (Character Error Rate) columns showing
    how each OCR reading compares to the closest ground truth plate.
    
    Returns a BytesIO binary buffer containing the spreadsheet so the web dashboard can stream it directly.
    """
    if ground_truth_plates is None:
        ground_truth_plates = []
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ALPR Detections"
    
    # 1. Define columns headers. Columns G and H are reserved for rendering embedded images.
    #    Columns I and J are for CER metrics when ground truth is available.
    headers = ["Track ID", "Timestamp", "Vehicle Type", "Color", "Plate Number", "Confidence", "Vehicle Image", "Plate Crop"]
    if ground_truth_plates:
        headers.extend(["Matched GT", "CER"])
    worksheet.append(headers)
    
    # 2. Styling the header row to look clean and professional (white text on a deep blue background)
    # Color hex '1F497D' is a professional dark steel blue
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # Apply the styling to all cells in the first row
    for column_index in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Set header row height to 25px to give it some comfortable breathing room
    worksheet.row_dimensions[1].height = 25
    
    # Stylings for standard data cells
    font_regular = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    
    # 3. Populate rows with vehicle telemetry data
    for idx, row in detections_df.iterrows():
        # Excel is 1-indexed, and row 1 is the header. So data begins at index + 2.
        row_index = idx + 2
        
        track_id = row.get("track_id", "")
        timestamp = row.get("timestamp", "")
        vehicle_type = row.get("vehicle_type", "")
        color = row.get("color", "")
        plate_number = row.get("plate_number", "")
        confidence = row.get("confidence", 0.0)
        snapshot_path = row.get("snapshot_path", "")
        
        # Populate text cells, centering values
        worksheet.cell(row=row_index, column=1, value=int(track_id) if str(track_id).isdigit() else track_id)
        worksheet.cell(row=row_index, column=2, value=str(timestamp))
        worksheet.cell(row=row_index, column=3, value=str(vehicle_type).capitalize())
        worksheet.cell(row=row_index, column=4, value=str(color).capitalize() if pd.notna(color) else "")
        worksheet.cell(row=row_index, column=5, value=str(plate_number))
        
        # Format confidence cell to display as a percentage (e.g. 0.825 becomes "83%")
        confidence_value = float(confidence) if pd.notna(confidence) else 0.0
        confidence_cell = worksheet.cell(row=row_index, column=6, value=confidence_value)
        confidence_cell.number_format = "0%"
        
        # Apply standard font styling to the text columns (columns 1 to 6)
        for column_index in range(1, 7):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = font_regular
            cell.alignment = align_center
        
        # 3b. Compute and populate CER columns if ground truth is available
        if ground_truth_plates and plate_number:
            best_gt, best_cer = find_best_ground_truth_match(str(plate_number), ground_truth_plates)
            
            gt_cell = worksheet.cell(row=row_index, column=9, value=str(best_gt) if best_gt else "No Match")
            gt_cell.font = font_regular
            gt_cell.alignment = align_center
            
            if best_cer is not None:
                cer_cell = worksheet.cell(row=row_index, column=10, value=best_cer)
                cer_cell.number_format = "0.00%"
                cer_cell.font = font_regular
                cer_cell.alignment = align_center
                
                # Color-code CER cells: green for perfect, yellow for moderate, red for bad
                if best_cer == 0.0:
                    cer_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif best_cer <= 0.3:
                    cer_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cer_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                worksheet.cell(row=row_index, column=10, value="N/A").font = font_regular
        elif ground_truth_plates:
            worksheet.cell(row=row_index, column=9, value="").font = font_regular
            worksheet.cell(row=row_index, column=10, value="").font = font_regular
        
        # Set height of each data row to 80px to accommodate the vehicle snapshots and plate crops
        worksheet.row_dimensions[row_index].height = 80
        
        # 4. Locate and embed corresponding processed plate crop for this track
        crop_glob = f"outputs/plate_crops/Processed/*_track{track_id}_processed.jpg"
        matching_files = glob.glob(crop_glob)
        if matching_files:
            crop_path = matching_files[0]
            if os.path.exists(crop_path):
                try:
                    img_plate = OpenpyxlImage(crop_path)
                    # Constrain image size to 100x35 pixels so it fits cleanly in column H
                    img_plate.width = 100
                    img_plate.height = 35
                    worksheet.add_image(img_plate, f"H{row_index}")
                except Exception:
                    worksheet.cell(row=row_index, column=8, value="Img Err")
        else:
            worksheet.cell(row=row_index, column=8, value="No Crop")
 
        # 5. Locate and embed full vehicle context snapshot image
        if snapshot_path and os.path.exists(snapshot_path):
            try:
                img_vehicle = OpenpyxlImage(snapshot_path)
                # Constrain image size to 120x70 pixels so it fits cleanly inside column G
                img_vehicle.width = 120
                img_vehicle.height = 70
                worksheet.add_image(img_vehicle, f"G{row_index}")
            except Exception:
                worksheet.cell(row=row_index, column=7, value="Img Err")
        else:
            worksheet.cell(row=row_index, column=7, value="No Image")
            
    # 6. Auto-fit column widths based on text length to prevent text clipping
    for column in worksheet.columns:
        max_len = 0
        column_letter = get_column_letter(column[0].column)
        
        # Column G is reserved for the vehicle image; keep width at 20
        if column_letter == "G":
            worksheet.column_dimensions["G"].width = 20
            continue
        # Column H is reserved for the plate crop image; keep width at 18
        if column_letter == "H":
            worksheet.column_dimensions["H"].width = 18
            continue
            
        # For all text columns, find the longest string value
        for cell in column:
            value_string = str(cell.value or '')
            if len(value_string) > max_len:
                max_len = len(value_string)
        # Add 3 characters padding; set a minimum default width of 10
        worksheet.column_dimensions[column_letter].width = max(max_len + 3, 10)
    
    # 7. Create a CER Summary sheet if ground truth plates were provided
    if ground_truth_plates:
        detections_list = detections_df.to_dict(orient="records")
        cer_summary = compute_average_cer(detections_list, ground_truth_plates)
        
        summary_ws = workbook.create_sheet(title="CER Summary")
        
        # Summary header styling
        summary_header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        summary_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        summary_font = Font(name="Arial", size=11)
        summary_bold = Font(name="Arial", size=11, bold=True)
        
        # Title row
        title_cell = summary_ws.cell(row=1, column=1, value="CER Evaluation Summary")
        title_cell.font = Font(name="Arial", size=14, bold=True)
        summary_ws.merge_cells("A1:C1")
        
        # Overall metrics
        summary_ws.cell(row=3, column=1, value="Average CER:").font = summary_bold
        avg_cer = cer_summary.get("average_cer")
        avg_cell = summary_ws.cell(row=3, column=2, value=avg_cer if avg_cer is not None else "N/A")
        if isinstance(avg_cer, (int, float)):
            avg_cell.number_format = "0.00%"
        avg_cell.font = summary_font
        
        summary_ws.cell(row=4, column=1, value="Matched Detections:").font = summary_bold
        summary_ws.cell(row=4, column=2, value=cer_summary.get("matched_count", 0)).font = summary_font
        
        summary_ws.cell(row=5, column=1, value="Total Detections:").font = summary_bold
        summary_ws.cell(row=5, column=2, value=cer_summary.get("total_detections", 0)).font = summary_font
        
        summary_ws.cell(row=6, column=1, value="Ground Truth Plates:").font = summary_bold
        summary_ws.cell(row=6, column=2, value=len(ground_truth_plates)).font = summary_font
        
        # Ground truth plates list
        summary_ws.cell(row=8, column=1, value="Ground Truth Plates").font = summary_bold
        for gt_idx, gt_plate in enumerate(ground_truth_plates):
            summary_ws.cell(row=9 + gt_idx, column=1, value=gt_plate).font = summary_font
        
        # Per-detection CER breakdown table
        breakdown_start = 9 + len(ground_truth_plates) + 1
        bd_headers = ["OCR Reading", "Matched GT", "CER"]
        for col_idx, hdr in enumerate(bd_headers, 1):
            cell = summary_ws.cell(row=breakdown_start, column=col_idx, value=hdr)
            cell.font = summary_header_font
            cell.fill = summary_header_fill
            cell.alignment = align_center
        
        for det_idx, det in enumerate(cer_summary.get("per_detection", [])):
            r = breakdown_start + 1 + det_idx
            summary_ws.cell(row=r, column=1, value=det.get("plate_number", "")).font = summary_font
            summary_ws.cell(row=r, column=2, value=det.get("matched_ground_truth", "")).font = summary_font
            cer_val = det.get("cer")
            cer_cell = summary_ws.cell(row=r, column=3, value=cer_val if cer_val is not None else "N/A")
            if isinstance(cer_val, (int, float)):
                cer_cell.number_format = "0.00%"
            cer_cell.font = summary_font
            cer_cell.alignment = align_center
        
        # Auto-fit summary columns
        for col_idx in range(1, 4):
            summary_ws.column_dimensions[get_column_letter(col_idx)].width = 22
        
    # Save the Excel file in-memory and rewind the stream position to the beginning
    out_buf = BytesIO()
    workbook.save(out_buf)
    out_buf.seek(0)
    return out_buf
